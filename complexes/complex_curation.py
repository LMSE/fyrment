
from pandas import ExcelWriter
from openpyxl import load_workbook
import urllib
import pandas as pd

def process_fog_annotation(oid,fog_annotation):
	"""
	Could and should clean this up
	Purpose is to take in a fog_annotation in FYRMENT and oid
	Check if all essential subunits are present
	"""
	def process_complex_variants(oid,fog_annotation):
		complex_check={}
		complex_variants=fog_annotation.split('||')
		for i,complex_variant in enumerate(complex_variants):
			complex_check[i]=[]
			subcomplexes=complex_variant.split('&&')
			for subcomplex in subcomplexes:
				subcomplex_list=subcomplex.replace('(','').replace(')','').replace(' ','').split('|')
				for subcomplex_component in subcomplex_list:
					fogs=subcomplex_component.split('&')
					for fog in fogs:
						complex_check[i].append(fog)
		complex_include={}
		for i in sorted(complex_check.keys()):
			# list of essential FOG's
			complex_essential=[fog[:8] for fog in complex_check[i] if '?' not in fog]
			# list of essential FOG's in oid
			complex_esssential_oid=[aybrah[oid][fog] for fog in complex_essential if len(aybrah[oid][fog[:8]])>0 ]
			# bypass if there is no oid (pan-model)
			#if len(inclusion_oids)>1:
			#	complex_include[i]=complex_check[i]
			# check if organism has all the designated essential proteins
			if len(complex_esssential_oid)==len(complex_essential):
				#oid=inclusion_oids[0]
				fogs_processed=[fog[:8] for fog in complex_check[i] if len(aybrah[oid][fog[:8]])>0]
				if len(fogs_processed)>0:
					complex_include[i]=fogs_processed
				print oid,'include complex '+str(i)
			else:
				# return empty
				print 'oid','not all subunits for complex '+str(i)
				#complex_include=None
		return complex_include
	# fog_annotation requires parsing
	if '&' in fog_annotation:
		# complex that meet essentialy requir.
		complex_include=process_complex_variants(oid,fog_annotation)
		fogs=complex_include
		#if oid != None:
		#	fogs=sorted(set([fog[:8] for key in complex_include for fog in complex_include[key] if len(aybra[oid][fog[:8]>0])]))
	# simple or fog_annotation
	else:
		fogs=filter(None,fog_annotation.replace('(','').replace(')','').replace(' ','').split('|'))
		if oid==None:
			fogs={i:[fog] for i,fog in enumerate(fogs)}
		else:
			fogs={i:[fog] for i,fog in enumerate(fogs) if len(aybrah[oid][fog])>0}
	return fogs



root_aybrah='../aybrah/'



aybrah=pd.read_csv(root_aybrah+'aybrah.tsv',sep='\t').fillna('').set_index('FOG')

version_aybrah=urllib.urlopen(root_aybrah+'version.txt').read()

path_aybrah_xlsx=root_aybrah+'aybrah.xlsx'
organisms=pd.read_excel(path_aybrah_xlsx,sheet_name='curated_taxonomy_fungi').fillna('')
indices_to_drop=organisms[organisms.oid.str.contains('ani|cpr')].index.tolist()
organisms=organisms.drop(indices_to_drop)


taxonomy=pd.read_excel(path_aybrah_xlsx,sheet_name='taxon_nodes').fillna('')
#taxonomy.drop(46)
drop_these=[index for oid in ['cpr','ani'] for index in taxonomy[taxonomy['oids']==oid].index.tolist()]
taxonomy=taxonomy.drop(drop_these)


# need to make sure ? fogs are captured

path_fyrment_xlsx='fyrment.xlsx'
rxns=pd.read_excel(path_fyrment_xlsx,sheet_name='reactions',encoding='ascii',skiprows=[0]).fillna('').set_index('Rxn name')


writer = ExcelWriter('./complexes/fyrment_complex_curation.xlsx')

fog_annotations=list(set(rxns[rxns.FOG.str.contains('&')].FOG.tolist()))
for number,fog_annotation in enumerate(fog_annotations):
	rxnids=';'.join(rxns[rxns.FOG==fog_annotation].index.tolist())
	formula=';'.join(rxns[rxns.FOG==fog_annotation].Formula.tolist())
	complex_check={}
	complex_variants=fog_annotation.split('||')
	for i,complex_variant in enumerate(complex_variants):
		complex_check[i]=[]
		subcomplexes=complex_variant.split('&&')
		for subcomplex in subcomplexes:
			subcomplex_list=subcomplex.replace('(','').replace(')','').replace(' ','').split('|')
			for subcomplex_component in subcomplex_list:
				fogs=subcomplex_component.split('&')
				for fog in fogs:
					complex_check[i].append(fog)
	fogs=[fog[:8] for i in complex_check.keys() for fog in complex_check[i]]
	fogs_nonessential=[fog[:8] for fog in fogs if '?' in fog]
	columns=['Rxn name','Formula','Include']+[(i,str(fog)) for i in complex_check.keys() for fog in complex_check[i]]
	df_temp=pd.DataFrame(columns=columns)
	for order,row in organisms.iterrows():
		oid=organisms['oid'][order]
		complexes_parsed=process_fog_annotation(oid,fog_annotation)
		include=len(complexes_parsed)
		row=[rxnids,formula,include]+[0]*(len(columns)-3)
		df_temp.loc[oid]=row
		for num,fog in df_temp.columns[3:].tolist():
			fog
			seqids=aybrah[oid][fog[:8]]
			if len(seqids)>0:
				df_temp.loc[oid,(num,fog)]=1
	df_temp.to_excel(writer,str(number))





writer.save()

"""
update colours to make it easier to identify missing subunits

"""

from openpyxl.styles import Color, PatternFill, Font, Border


yellowFill = PatternFill(start_color='FCFF00',
                   end_color='FCFF00',
                   fill_type='solid')

greyFill = PatternFill(start_color='a5a5a5',
                   end_color='a5a5a5',
                   fill_type='solid')




wb = load_workbook(filename = './complexes/fyrment_complex_curation.xlsx' )


for sheetname in wb.sheetnames:
	# highlight non-essential
	ws=wb[sheetname]
	for col in ws.iter_cols(min_row=1,max_row=1,min_col=2):
		for cell in col:
			if '?' in cell.value:
				cell.fill =yellowFill
	# grey out missing genes
	for col in ws.iter_cols(min_row=2):
		for cell in col:
			if str(cell.value)=='0':
				cell.fill = greyFill


wb.save('./complexes/fyrment_complex_curation.xlsx')


# https://stackoverflow.com/questions/30484220/python-fill-cells-with-colors-using-openpyxl
# ws['A1'].fill = redFill should work fine.

"""
	mismatching_rxn_columns=[cell.value \
			for header,col in zip(rxn_headers,ws.iter_cols(min_row=1, max_col=15, max_row=1)) \
			for cell in col if header != cell.value]
"""

"""
complex_variants_previous = pd.ExcelFile('./complexes/fyrment_complex_curation_prev.xlsx')
complex_variants_current = pd.ExcelFile('./complexes/fyrment_complex_curation.xlsx')


complex_variants_current.sheet_names

df_prev={sheet:pd.read_excel('./complexes/fyrment_complex_curation_prev.xlsx',sheetname=sheet) for sheet in complex_variants_previous.sheet_names}

df_current={sheet:pd.read_excel('./complexes/fyrment_complex_curation.xlsx',sheetname=sheet) for sheet in complex_variants_current.sheet_names}

def matching_complex(rxns_prev):
	for sheet_current in complex_variants_current.sheet_names:
		rxns_current=set(df_current[sheet_current]['Rxn name']['rgm'].split(';'))
		if rxns_current==rxns_prev:
			return sheet_current


sheets_prev=complex_variants_previous.sheet_names


while len(sheets_prev)>0:
	sheet_prev=sheets_prev.pop(0)
	#break
	rxns_prev=set(df_prev[sheet_prev]['Rxn name']['rgm'].split(';'))
	sheet_current=matching_complex(rxns_prev)
	#
	oids_prev=df_prev[sheet_prev][df_prev[sheet_prev]['Include']>0].index.tolist()
	if sheet_current==None:
		oids_current=[]
	else:
		oids_current=df_current[sheet_current][df_current[sheet_current]['Include']>0].index.tolist()
	if set(oids_prev)!=set(oids_current):
		print(rxns_prev)
		sheet_prev,sheet_current
		set(oids_current)-set(oids_prev)
		set(oids_prev)-set(oids_current)
		print('\n\n\n')


"""